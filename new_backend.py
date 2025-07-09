"""
Enhanced Therapeutic Companion Backend
With PostgreSQL, Authentication, Role-Based Access, and Client Reports
"""
import random
import string
from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS
from flask_bcrypt import Bcrypt
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from functools import wraps
import os
from pathlib import Path
import secrets
import jwt
from datetime import datetime, timedelta, date
from sqlalchemy import and_, or_, func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO

# Create Flask app
app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
app.static_folder = BASE_DIR
app.template_folder = BASE_DIR

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://localhost/therapy_companion'
).replace('postgres://', 'postgresql://')  # Fix for Render
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('PRODUCTION', False)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('SMTP_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('SYSTEM_EMAIL')
app.config['MAIL_PASSWORD'] = os.environ.get('SYSTEM_EMAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('SYSTEM_EMAIL')


# Database connection pooling - ADD THIS SECTION
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 5,
    'pool_recycle': 300,  # Recycle connections after 5 minutes
    'pool_pre_ping': True,  # Test connections before using them
    'max_overflow': 10,
    'connect_args': {
        'connect_timeout': 10,
        'options': '-c statement_timeout=30000'  # 30 second statement timeout
    }
}




# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)
bcrypt = Bcrypt(app)
CORS(app, supports_credentials=True)

# JWT configuration
JWT_SECRET = app.config['SECRET_KEY']
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24


# ============= DATABASE MODELS =============

class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)

    # Relationships
    therapist = db.relationship('Therapist', backref='user', uselist=False, cascade='all, delete-orphan')
    client = db.relationship('Client', backref='user', uselist=False, cascade='all, delete-orphan')
    session_tokens = db.relationship('SessionToken', backref='user', cascade='all, delete-orphan')


class Therapist(db.Model):
    __tablename__ = 'therapists'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True)
    license_number = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    organization = db.Column(db.String(255))
    specializations = db.Column(db.JSON)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    clients = db.relationship('Client', backref='therapist', lazy='dynamic')
    notes = db.relationship('TherapistNote', backref='therapist', lazy='dynamic')


class Client(db.Model):
    __tablename__ = 'clients'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True)
    client_serial = db.Column(db.String(50), unique=True, nullable=False)
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    start_date = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    checkins = db.relationship('DailyCheckin', backref='client', lazy='dynamic', cascade='all, delete-orphan')
    tracking_plans = db.relationship('ClientTrackingPlan', backref='client', lazy='dynamic',
                                     cascade='all, delete-orphan')
    goals = db.relationship('WeeklyGoal', backref='client', lazy='dynamic', cascade='all, delete-orphan')
    reminders = db.relationship('Reminder', backref='client', lazy='dynamic', cascade='all, delete-orphan')


class TrackingCategory(db.Model):
    __tablename__ = 'tracking_categories'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    scale_min = db.Column(db.Integer, default=1)
    scale_max = db.Column(db.Integer, default=5)
    is_default = db.Column(db.Boolean, default=False)


class ClientTrackingPlan(db.Model):
    __tablename__ = 'client_tracking_plans'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    category_id = db.Column(db.Integer, db.ForeignKey('tracking_categories.id'))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    category = db.relationship('TrackingCategory', backref='client_plans')


class WeeklyGoal(db.Model):
    __tablename__ = 'weekly_goals'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    goal_text = db.Column(db.Text, nullable=False)
    week_start = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    completions = db.relationship('GoalCompletion', backref='goal', lazy='dynamic', cascade='all, delete-orphan')


class DailyCheckin(db.Model):
    __tablename__ = 'daily_checkins'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    checkin_date = db.Column(db.Date, nullable=False)
    checkin_time = db.Column(db.Time, nullable=False)
    emotional_value = db.Column(db.Integer)
    emotional_notes = db.Column(db.Text)
    medication_value = db.Column(db.Integer)
    medication_notes = db.Column(db.Text)
    activity_value = db.Column(db.Integer)
    activity_notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint('client_id', 'checkin_date'),)


class CategoryResponse(db.Model):
    __tablename__ = 'category_responses'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    category_id = db.Column(db.Integer, db.ForeignKey('tracking_categories.id'))
    response_date = db.Column(db.Date, nullable=False)
    value = db.Column(db.Integer, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    category = db.relationship('TrackingCategory', backref='responses')


class GoalCompletion(db.Model):
    __tablename__ = 'goal_completions'

    id = db.Column(db.Integer, primary_key=True)
    goal_id = db.Column(db.Integer, db.ForeignKey('weekly_goals.id'))
    completion_date = db.Column(db.Date, nullable=False)
    completed = db.Column(db.Boolean, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint('goal_id', 'completion_date'),)


class Reminder(db.Model):
    __tablename__ = 'reminders'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    reminder_type = db.Column(db.String(50), nullable=False)
    reminder_time = db.Column(db.Time, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    last_sent = db.Column(db.DateTime)


class Report(db.Model):
    __tablename__ = 'reports'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    report_type = db.Column(db.String(50), nullable=False)
    week_start = db.Column(db.Date, nullable=False)
    generated_at = db.Column(db.DateTime, default=datetime.utcnow)
    file_path = db.Column(db.Text)
    data = db.Column(db.JSON)


class TherapistNote(db.Model):
    __tablename__ = 'therapist_notes'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    note_type = db.Column(db.String(50), default='general')
    content = db.Column(db.Text, nullable=False)
    is_mission = db.Column(db.Boolean, default=False)
    mission_completed = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime)


class SessionToken(db.Model):
    __tablename__ = 'session_tokens'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    token = db.Column(db.String(255), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


# ============= AUTHENTICATION HELPERS =============

def generate_token(user_id, role):
    """Generate JWT token"""
    payload = {
        'user_id': user_id,
        'role': role,
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def verify_token(token):
    """Verify JWT token"""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def require_auth(allowed_roles=None):
    """Authentication decorator"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            auth_header = request.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                return jsonify({'error': 'Invalid authorization header'}), 401

            token = auth_header.replace('Bearer ', '')
            payload = verify_token(token)

            if not payload:
                return jsonify({'error': 'Invalid or expired token'}), 401

            # Check if user exists and is active
            user = User.query.get(payload['user_id'])
            if not user or not user.is_active:
                return jsonify({'error': 'User not found or inactive'}), 401

            # Check role permissions
            if allowed_roles and user.role not in allowed_roles:
                return jsonify({'error': 'Insufficient permissions'}), 403

            # Add user info to request
            request.current_user = user
            request.user_id = user.id
            request.user_role = user.role

            return f(*args, **kwargs)

        return decorated_function

    return decorator


def generate_client_serial():
    """Generate unique client serial number"""
    import random
    import string
    while True:
        serial = 'C' + ''.join(random.choices(string.digits, k=8))
        if not Client.query.filter_by(client_serial=serial).first():
            return serial


# ============= HTML PAGE ROUTES =============

@app.route('/')
def index():
    """Serve the main HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'index.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            app.logger.error(f"index.html not found at {file_path}")
            return "index.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving index.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/login.html')
def login_page():
    """Serve the login HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'login.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            app.logger.error(f"login.html not found at {file_path}")
            return "login.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving login.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/therapist-dashboard.html')
def therapist_dashboard_page():
    """Serve the therapist dashboard HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'therapist_dashboard.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            # Try alternative filename
            alt_path = os.path.join(BASE_DIR, 'therapist-dashboard.html')
            if os.path.exists(alt_path):
                return send_file(alt_path)
            app.logger.error(f"therapist_dashboard.html not found at {file_path}")
            return "therapist_dashboard.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving therapist_dashboard.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/client-dashboard.html')
def client_dashboard_page():
    """Serve the client dashboard HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'client_dashboard.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            # Try alternative filename
            alt_path = os.path.join(BASE_DIR, 'client-dashboard.html')
            if os.path.exists(alt_path):
                return send_file(alt_path)
            app.logger.error(f"client_dashboard.html not found at {file_path}")
            return "client_dashboard.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving client_dashboard.html: {e}")
        return f"Error: {str(e)}", 500


# ============= API ENDPOINTS =============

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register new user (therapist or client)"""
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')
        role = data.get('role')

        # Validate input
        if not all([email, password, role]):
            return jsonify({'error': 'Missing required fields'}), 400

        if role not in ['therapist', 'client']:
            return jsonify({'error': 'Invalid role'}), 400

        # Check if email exists
        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already registered'}), 400

        # Create user
        user = User(
            email=email,
            password_hash=bcrypt.generate_password_hash(password).decode('utf-8'),
            role=role
        )
        db.session.add(user)
        db.session.flush()

        # Create role-specific profile
        if role == 'therapist':
            therapist = Therapist(
                user_id=user.id,
                license_number=data.get('license_number', ''),
                name=data.get('name', ''),
                organization=data.get('organization', ''),
                specializations=data.get('specializations', [])
            )
            db.session.add(therapist)
        else:  # client
            client = Client(
                user_id=user.id,
                client_serial=generate_client_serial(),
                therapist_id=data.get('therapist_id'),
                start_date=date.today()
            )
            db.session.add(client)

            # Add default tracking categories
            default_categories = TrackingCategory.query.filter_by(is_default=True).all()
            for category in default_categories:
                plan = ClientTrackingPlan(
                    client_id=client.id,
                    category_id=category.id
                )
                db.session.add(plan)

        db.session.commit()

        # Generate token
        token = generate_token(user.id, role)

        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'id': user.id,
                'email': user.email,
                'role': user.role
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login user"""
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')

        if not all([email, password]):
            return jsonify({'error': 'Missing email or password'}), 400

        # Find user
        user = User.query.filter_by(email=email).first()
        if not user or not bcrypt.check_password_hash(user.password_hash, password):
            return jsonify({'error': 'Invalid credentials'}), 401

        if not user.is_active:
            return jsonify({'error': 'Account deactivated'}), 401

        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Generate token
        token = generate_token(user.id, user.role)

        # Get role-specific data
        response_data = {
            'success': True,
            'token': token,
            'user': {
                'id': user.id,
                'email': user.email,
                'role': user.role
            }
        }

        if user.role == 'therapist' and user.therapist:
            response_data['therapist'] = {
                'id': user.therapist.id,
                'name': user.therapist.name,
                'license_number': user.therapist.license_number,
                'organization': user.therapist.organization
            }
        elif user.role == 'client' and user.client:
            response_data['client'] = {
                'id': user.client.id,
                'serial': user.client.client_serial,
                'start_date': user.client.start_date.isoformat()
            }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= THERAPIST ENDPOINTS =============

@app.route('/api/therapist/dashboard', methods=['GET'])
@require_auth(['therapist'])
def therapist_dashboard():
    """Get therapist dashboard data"""
    try:
        therapist = request.current_user.therapist

        # Get client statistics
        total_clients = therapist.clients.count()
        active_clients = therapist.clients.filter_by(is_active=True).count()

        # Get recent activity
        recent_checkins = db.session.query(DailyCheckin).join(Client).filter(
            Client.therapist_id == therapist.id,
            DailyCheckin.checkin_date >= date.today() - timedelta(days=7)
        ).count()

        # Get pending missions
        pending_missions = TherapistNote.query.filter_by(
            therapist_id=therapist.id,
            is_mission=True,
            mission_completed=False
        ).count()

        return jsonify({
            'success': True,
            'therapist': {
                'name': therapist.name,
                'license_number': therapist.license_number,
                'organization': therapist.organization
            },
            'statistics': {
                'total_clients': total_clients,
                'active_clients': active_clients,
                'recent_checkins': recent_checkins,
                'pending_missions': pending_missions
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/clients', methods=['GET'])
@require_auth(['therapist'])
def get_therapist_clients():
    """Get list of therapist's clients"""
    try:
        therapist = request.current_user.therapist

        # Get filter parameters
        status = request.args.get('status', 'all')
        sort_by = request.args.get('sort_by', 'start_date')

        # Build query
        query = therapist.clients

        if status == 'active':
            query = query.filter_by(is_active=True)
        elif status == 'inactive':
            query = query.filter_by(is_active=False)

        # Sort
        if sort_by == 'start_date':
            query = query.order_by(Client.start_date.desc())
        elif sort_by == 'serial':
            query = query.order_by(Client.client_serial)

        clients = query.all()

        # Build response
        client_data = []
        for client in clients:
            # Get last check-in
            last_checkin = client.checkins.order_by(DailyCheckin.checkin_date.desc()).first()

            # Get completion rate for last week
            week_start = date.today() - timedelta(days=date.today().weekday())
            week_checkins = client.checkins.filter(
                DailyCheckin.checkin_date >= week_start
            ).count()

            client_data.append({
                'id': client.id,
                'serial': client.client_serial,
                'start_date': client.start_date.isoformat(),
                'is_active': client.is_active,
                'last_checkin': last_checkin.checkin_date.isoformat() if last_checkin else None,
                'week_completion': f"{week_checkins}/7",
                'tracking_categories': [
                    plan.category.name for plan in client.tracking_plans.filter_by(is_active=True)
                ]
            })

        return jsonify({
            'success': True,
            'clients': client_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/client/<int:client_id>', methods=['GET'])
@require_auth(['therapist'])
def get_client_details(client_id):
    """Get detailed client information"""
    try:
        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Get tracking plans
        tracking_plans = []
        for plan in client.tracking_plans.filter_by(is_active=True):
            tracking_plans.append({
                'id': plan.id,
                'category': plan.category.name,
                'description': plan.category.description
            })

        # Get active goals
        active_goals = []
        week_start = date.today() - timedelta(days=date.today().weekday())
        for goal in client.goals.filter_by(
                week_start=week_start,
                is_active=True
        ):
            # Get completions for this week
            completions = {}
            for i in range(7):
                day = week_start + timedelta(days=i)
                completion = goal.completions.filter_by(completion_date=day).first()
                completions[day.isoformat()] = completion.completed if completion else None

            active_goals.append({
                'id': goal.id,
                'text': goal.goal_text,
                'completions': completions
            })

        # Get recent check-ins
        recent_checkins = []
        for checkin in client.checkins.order_by(
                DailyCheckin.checkin_date.desc()
        ).limit(7):
            recent_checkins.append({
                'date': checkin.checkin_date.isoformat(),
                'emotional': checkin.emotional_value,
                'medication': checkin.medication_value,
                'activity': checkin.activity_value
            })

        # Get therapist notes/missions
        notes = []
        for note in TherapistNote.query.filter_by(
                client_id=client_id,
                therapist_id=therapist.id
        ).order_by(TherapistNote.created_at.desc()).limit(10):
            notes.append({
                'id': note.id,
                'type': note.note_type,
                'content': note.content,
                'is_mission': note.is_mission,
                'completed': note.mission_completed,
                'created_at': note.created_at.isoformat()
            })

        return jsonify({
            'success': True,
            'client': {
                'id': client.id,
                'serial': client.client_serial,
                'start_date': client.start_date.isoformat(),
                'is_active': client.is_active,
                'tracking_plans': tracking_plans,
                'active_goals': active_goals,
                'recent_checkins': recent_checkins,
                'notes': notes
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/create-client', methods=['POST'])
@require_auth(['therapist'])
def create_client():
    """Create new client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        # Create user account for client
        email = data.get('email')
        password = data.get('password', secrets.token_urlsafe(8))

        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already registered'}), 400

        # Create user
        user = User(
            email=email,
            password_hash=bcrypt.generate_password_hash(password).decode('utf-8'),
            role='client'
        )
        db.session.add(user)
        db.session.flush()

        # Create client
        client = Client(
            user_id=user.id,
            client_serial=generate_client_serial(),
            therapist_id=therapist.id,
            start_date=date.today()
        )
        db.session.add(client)
        db.session.flush()

        # Add tracking categories
        category_ids = data.get('tracking_categories', [])
        if not category_ids:
            # Add default categories
            default_categories = TrackingCategory.query.filter_by(is_default=True).all()
            category_ids = [cat.id for cat in default_categories]

        for cat_id in category_ids:
            plan = ClientTrackingPlan(
                client_id=client.id,
                category_id=cat_id
            )
            db.session.add(plan)

        # Add initial goals if provided
        goals = data.get('initial_goals', [])
        week_start = date.today() - timedelta(days=date.today().weekday())
        for goal_text in goals:
            goal = WeeklyGoal(
                client_id=client.id,
                therapist_id=therapist.id,
                goal_text=goal_text,
                week_start=week_start
            )
            db.session.add(goal)

        # Add welcome note
        welcome_note = TherapistNote(
            client_id=client.id,
            therapist_id=therapist.id,
            note_type='welcome',
            content=f"Welcome to therapy! Your journey begins today. Your temporary password is: {password}"
        )
        db.session.add(welcome_note)

        db.session.commit()

        return jsonify({
            'success': True,
            'client': {
                'id': client.id,
                'serial': client.client_serial,
                'email': email,
                'temporary_password': password
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/add-goal', methods=['POST'])
@require_auth(['therapist'])
def add_weekly_goal():
    """Add weekly goal for client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        goal_text = data.get('goal_text')
        week_start_str = data.get('week_start')

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week start
        if week_start_str:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        else:
            # Default to current week
            week_start = date.today() - timedelta(days=date.today().weekday())

        # Create goal
        goal = WeeklyGoal(
            client_id=client_id,
            therapist_id=therapist.id,
            goal_text=goal_text,
            week_start=week_start
        )
        db.session.add(goal)
        db.session.commit()

        return jsonify({
            'success': True,
            'goal': {
                'id': goal.id,
                'text': goal.goal_text,
                'week_start': goal.week_start.isoformat()
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/add-note', methods=['POST'])
@require_auth(['therapist'])
def add_therapist_note():
    """Add note or mission for client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        content = data.get('content')
        is_mission = data.get('is_mission', False)
        note_type = data.get('note_type', 'general')

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Create note
        note = TherapistNote(
            client_id=client_id,
            therapist_id=therapist.id,
            content=content,
            is_mission=is_mission,
            note_type=note_type
        )
        db.session.add(note)
        db.session.commit()

        return jsonify({
            'success': True,
            'note': {
                'id': note.id,
                'content': note.content,
                'is_mission': note.is_mission,
                'created_at': note.created_at.isoformat()
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============= CLIENT ENDPOINTS =============

@app.route('/api/client/dashboard', methods=['GET'])
@require_auth(['client'])
def client_dashboard():
    """Get client dashboard data"""
    try:
        client = request.current_user.client

        # Get today's check-in status
        today_checkin = client.checkins.filter_by(checkin_date=date.today()).first()

        # Get active tracking categories
        tracking_categories = []
        for plan in client.tracking_plans.filter_by(is_active=True):
            # Get today's response
            today_response = CategoryResponse.query.filter_by(
                client_id=client.id,
                category_id=plan.category_id,
                response_date=date.today()
            ).first()

            tracking_categories.append({
                'id': plan.category_id,
                'name': plan.category.name,
                'description': plan.category.description,
                'today_value': today_response.value if today_response else None
            })

        # Get this week's goals
        week_start = date.today() - timedelta(days=date.today().weekday())
        weekly_goals = []
        for goal in client.goals.filter_by(
                week_start=week_start,
                is_active=True
        ):
            # Get today's completion
            today_completion = goal.completions.filter_by(
                completion_date=date.today()
            ).first()

            weekly_goals.append({
                'id': goal.id,
                'text': goal.goal_text,
                'today_completed': today_completion.completed if today_completion else None
            })

        # Get reminders
        reminders = []
        for reminder in client.reminders.filter_by(is_active=True):
            reminders.append({
                'type': reminder.reminder_type,
                'time': reminder.reminder_time.strftime('%H:%M')
            })

        return jsonify({
            'success': True,
            'client': {
                'serial': client.client_serial,
                'start_date': client.start_date.isoformat()
            },
            'today': {
                'has_checkin': today_checkin is not None,
                'date': date.today().isoformat()
            },
            'tracking_categories': tracking_categories,
            'weekly_goals': weekly_goals,
            'reminders': reminders
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/checkin', methods=['POST'])
@require_auth(['client'])
def submit_checkin():
    """Submit daily check-in"""
    try:
        client = request.current_user.client
        data = request.json

        checkin_date = data.get('date', date.today().isoformat())
        checkin_date = datetime.strptime(checkin_date, '%Y-%m-%d').date()

        # Check if check-in exists
        existing = client.checkins.filter_by(checkin_date=checkin_date).first()

        if existing:
            # Update existing
            existing.checkin_time = datetime.now().time()
            existing.emotional_value = data.get('emotional_value')
            existing.emotional_notes = data.get('emotional_notes')
            existing.medication_value = data.get('medication_value')
            existing.medication_notes = data.get('medication_notes')
            existing.activity_value = data.get('activity_value')
            existing.activity_notes = data.get('activity_notes')
        else:
            # Create new
            checkin = DailyCheckin(
                client_id=client.id,
                checkin_date=checkin_date,
                checkin_time=datetime.now().time(),
                emotional_value=data.get('emotional_value'),
                emotional_notes=data.get('emotional_notes'),
                medication_value=data.get('medication_value'),
                medication_notes=data.get('medication_notes'),
                activity_value=data.get('activity_value'),
                activity_notes=data.get('activity_notes')
            )
            db.session.add(checkin)

        # Save category responses
        category_responses = data.get('category_responses', {})
        for cat_id, value in category_responses.items():
            # Check if response exists
            existing_response = CategoryResponse.query.filter_by(
                client_id=client.id,
                category_id=int(cat_id),
                response_date=checkin_date
            ).first()

            if existing_response:
                existing_response.value = value
            else:
                response = CategoryResponse(
                    client_id=client.id,
                    category_id=int(cat_id),
                    response_date=checkin_date,
                    value=value
                )
                db.session.add(response)

        # Save goal completions
        goal_completions = data.get('goal_completions', {})
        for goal_id, completed in goal_completions.items():
            # Check if completion exists
            existing_completion = GoalCompletion.query.filter_by(
                goal_id=int(goal_id),
                completion_date=checkin_date
            ).first()

            if existing_completion:
                existing_completion.completed = completed
            else:
                completion = GoalCompletion(
                    goal_id=int(goal_id),
                    completion_date=checkin_date,
                    completed=completed
                )
                db.session.add(completion)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Check-in saved successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/progress', methods=['GET'])
@require_auth(['client'])
def get_client_progress():
    """Get client's progress data"""
    try:
        client = request.current_user.client

        # Get date range
        end_date = date.today()
        start_date = end_date - timedelta(days=30)

        # Get check-ins
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(start_date, end_date)
        ).order_by(DailyCheckin.checkin_date).all()

        checkin_data = []
        for checkin in checkins:
            checkin_data.append({
                'date': checkin.checkin_date.isoformat(),
                'emotional': checkin.emotional_value,
                'medication': checkin.medication_value,
                'activity': checkin.activity_value
            })

        # Get category responses
        category_data = {}
        for plan in client.tracking_plans.filter_by(is_active=True):
            responses = CategoryResponse.query.filter(
                CategoryResponse.client_id == client.id,
                CategoryResponse.category_id == plan.category_id,
                CategoryResponse.response_date.between(start_date, end_date)
            ).order_by(CategoryResponse.response_date).all()

            category_data[plan.category.name] = [
                {
                    'date': resp.response_date.isoformat(),
                    'value': resp.value
                } for resp in responses
            ]

        return jsonify({
            'success': True,
            'progress': {
                'checkins': checkin_data,
                'categories': category_data
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= CLIENT REPORT ENDPOINTS =============

@app.route('/api/client/generate-report/<week>', methods=['GET'])
@require_auth(['client'])
def client_generate_report(week):
    """Generate client's own weekly report"""
    try:
        client = request.current_user.client

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Week {week_num} Report"

        # Header styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Cell styles
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"Weekly Progress Report - Week {week_num}, {year}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = header_alignment

        # Client info
        ws['A3'] = "Client ID:"
        ws['B3'] = client.client_serial
        ws['A4'] = "Week:"
        ws['B4'] = f"{week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')}"
        ws['A5'] = "Generated:"
        ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Headers
        headers = ['Date', 'Day', 'Time', 'Emotional (1-5)', 'Emotional Notes',
                   'Medication', 'Medication Notes', 'Activity (1-5)', 'Activity Notes']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Get check-ins for the week
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start.date(), week_end.date())
        ).order_by(DailyCheckin.checkin_date).all()

        # Populate data
        row = 8
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        for i in range(7):
            current_date = week_start + timedelta(days=i)
            checkin = next((c for c in checkins if c.checkin_date == current_date.date()), None)

            ws.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
            ws.cell(row=row, column=2).value = days[i]

            if checkin:
                ws.cell(row=row, column=3).value = checkin.checkin_time.strftime('%H:%M')
                ws.cell(row=row, column=4).value = checkin.emotional_value
                ws.cell(row=row, column=5).value = checkin.emotional_notes or ''

                # Medication formatting
                med_cell = ws.cell(row=row, column=6)
                if checkin.medication_value == 0:
                    med_cell.value = "N/A"
                elif checkin.medication_value == 1:
                    med_cell.value = "No"
                elif checkin.medication_value == 3:
                    med_cell.value = "Partial"
                elif checkin.medication_value == 5:
                    med_cell.value = "Yes"

                ws.cell(row=row, column=7).value = checkin.medication_notes or ''
                ws.cell(row=row, column=8).value = checkin.activity_value
                ws.cell(row=row, column=9).value = checkin.activity_notes or ''

                # Color coding for ratings
                emotional_cell = ws.cell(row=row, column=4)
                if checkin.emotional_value:
                    if checkin.emotional_value >= 4:
                        emotional_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    elif checkin.emotional_value == 3:
                        emotional_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    else:
                        emotional_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                activity_cell = ws.cell(row=row, column=8)
                if checkin.activity_value:
                    if checkin.activity_value >= 4:
                        activity_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    elif checkin.activity_value == 3:
                        activity_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    else:
                        activity_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                ws.cell(row=row, column=3).value = "No check-in"
                ws.cell(row=row, column=3).font = Font(italic=True, color="999999")

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = cell_border

            row += 1

        # Summary section
        row += 2
        ws.cell(row=row, column=1).value = "WEEKLY SUMMARY"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)

        row += 1
        total_checkins = len(checkins)
        ws.cell(row=row, column=1).value = "Total Check-ins:"
        ws.cell(row=row, column=2).value = f"{total_checkins}/7"

        if total_checkins > 0:
            # Average emotional rating
            avg_emotional = sum(c.emotional_value for c in checkins if c.emotional_value) / total_checkins
            row += 1
            ws.cell(row=row, column=1).value = "Average Emotional Rating:"
            ws.cell(row=row, column=2).value = f"{avg_emotional:.1f}/5"

            # Medication adherence
            med_adherent = sum(1 for c in checkins if c.medication_value == 5)
            row += 1
            ws.cell(row=row, column=1).value = "Medication Adherence:"
            ws.cell(row=row, column=2).value = f"{med_adherent}/{total_checkins} days"

            # Average activity
            avg_activity = sum(c.activity_value for c in checkins if c.activity_value) / total_checkins
            row += 1
            ws.cell(row=row, column=1).value = "Average Activity Level:"
            ws.cell(row=row, column=2).value = f"{avg_activity:.1f}/5"

        # Goals section if any
        weekly_goals = client.goals.filter_by(week_start=week_start.date()).all()
        if weekly_goals:
            row += 2
            ws.cell(row=row, column=1).value = "WEEKLY GOALS"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)

            for goal in weekly_goals:
                row += 1
                ws.cell(row=row, column=1).value = f"• {goal.goal_text}"

                # Goal completions
                completions = goal.completions.filter(
                    GoalCompletion.completion_date.between(week_start.date(), week_end.date())
                ).all()
                completed_days = sum(1 for c in completions if c.completed)
                ws.cell(row=row, column=3).value = f"Completed: {completed_days}/7 days"

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        filename = f"my_therapy_report_{client.client_serial}_week_{week_num}_{year}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/email-report', methods=['POST'])
@require_auth(['client'])
def client_email_report():
    """Prepare email report for client to send to therapist"""
    try:
        client = request.current_user.client
        data = request.json
        week = data.get('week')

        if not week:
            return jsonify({'error': 'Week is required'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get therapist info
        therapist = client.therapist
        therapist_email = therapist.user.email if therapist and therapist.user else "therapist@example.com"
        therapist_name = therapist.name if therapist else "Therapist"

        # Get check-ins for the week
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start.date(), week_end.date())
        ).order_by(DailyCheckin.checkin_date).all()

        # Build email content
        subject = f"Weekly Therapy Report - {client.client_serial} - Week {week_num}, {year}"

        content = f"""Dear {therapist_name},

Here is my weekly progress report for {week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')}.

CLIENT: {client.client_serial}
WEEK: {week_num}, {year}
CHECK-INS COMPLETED: {len(checkins)}/7

DAILY CHECK-IN SUMMARY:
"""

        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        for i in range(7):
            current_date = week_start + timedelta(days=i)
            checkin = next((c for c in checkins if c.checkin_date == current_date.date()), None)

            content += f"\n{days[i]} ({current_date.strftime('%m/%d')}):\n"

            if checkin:
                content += f"  ✓ Checked in at {checkin.checkin_time.strftime('%H:%M')}\n"
                content += f"  - Emotional State: {checkin.emotional_value}/5"
                if checkin.emotional_notes:
                    content += f" (Notes: {checkin.emotional_notes})"
                content += "\n"

                content += f"  - Medication: "
                if checkin.medication_value == 0:
                    content += "N/A"
                elif checkin.medication_value == 1:
                    content += "Not taken"
                elif checkin.medication_value == 3:
                    content += "Partially taken"
                elif checkin.medication_value == 5:
                    content += "Taken as prescribed"
                if checkin.medication_notes:
                    content += f" (Notes: {checkin.medication_notes})"
                content += "\n"

                content += f"  - Physical Activity: {checkin.activity_value}/5"
                if checkin.activity_notes:
                    content += f" (Notes: {checkin.activity_notes})"
                content += "\n"
            else:
                content += "  ✗ No check-in recorded\n"

        # Add summary
        if checkins:
            content += "\nWEEKLY SUMMARY:\n"

            total_checkins = len(checkins)
            avg_emotional = sum(c.emotional_value for c in checkins if c.emotional_value) / total_checkins
            med_adherent = sum(1 for c in checkins if c.medication_value == 5)
            avg_activity = sum(c.activity_value for c in checkins if c.activity_value) / total_checkins

            content += f"- Completion Rate: {total_checkins}/7 days ({(total_checkins / 7) * 100:.0f}%)\n"
            content += f"- Average Emotional Rating: {avg_emotional:.1f}/5\n"
            content += f"- Medication Adherence: {med_adherent}/{total_checkins} days\n"
            content += f"- Average Activity Level: {avg_activity:.1f}/5\n"

        # Add goals if any
        weekly_goals = client.goals.filter_by(week_start=week_start.date()).all()
        if weekly_goals:
            content += "\nWEEKLY GOALS:\n"
            for goal in weekly_goals:
                completions = goal.completions.filter(
                    GoalCompletion.completion_date.between(week_start.date(), week_end.date())
                ).all()
                completed_days = sum(1 for c in completions if c.completed)
                content += f"- {goal.goal_text}: Completed {completed_days}/7 days\n"

        content += f"\nReport generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        content += "\nBest regards,\n"
        content += f"Client {client.client_serial}"

        return jsonify({
            'success': True,
            'recipient': therapist_email,
            'subject': subject,
            'content': content,
            'note': 'Copy this email content to send to your therapist'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/week-checkins/<week>', methods=['GET'])
@require_auth(['client'])
def get_client_week_checkins(week):
    """Get client's check-ins for a specific week"""
    try:
        client = request.current_user.client

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get check-ins
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start.date(), week_end.date())
        ).all()

        # Format response
        checkin_data = {}
        for checkin in checkins:
            checkin_data[checkin.checkin_date.isoformat()] = {
                'time': checkin.checkin_time.strftime('%H:%M'),
                'emotional': checkin.emotional_value,
                'emotional_notes': checkin.emotional_notes,
                'medication': checkin.medication_value,
                'medication_notes': checkin.medication_notes,
                'activity': checkin.activity_value,
                'activity_notes': checkin.activity_notes
            }

        return jsonify({
            'success': True,
            'week_start': week_start.date().isoformat(),
            'week_end': week_end.date().isoformat(),
            'checkins': checkin_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/goals/<week>', methods=['GET'])
@require_auth(['client'])
def get_client_week_goals(week):
    """Get client's goals for a specific week"""
    try:
        client = request.current_user.client

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week start
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)

        # Get goals for the week
        goals = client.goals.filter_by(
            week_start=week_start.date(),
            is_active=True
        ).all()

        # Format response
        goals_data = []
        for goal in goals:
            # Get completions for the week
            completions = {}
            for i in range(7):
                day = week_start.date() + timedelta(days=i)
                completion = goal.completions.filter_by(completion_date=day).first()
                completions[day.isoformat()] = completion.completed if completion else None

            goals_data.append({
                'id': goal.id,
                'text': goal.goal_text,
                'week_start': goal.week_start.isoformat(),
                'completions': completions
            })

        return jsonify({
            'success': True,
            'goals': goals_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def create_weekly_report_excel(client, therapist, week_start, week_end, week_num, year):
    """Create Excel workbook for weekly report - shared by generate_report and email_therapy_report"""
    # Create Excel workbook
    wb = openpyxl.Workbook()

    # 1. Daily Check-ins Sheet
    ws_checkins = wb.active
    ws_checkins.title = "Daily Check-ins"

    # Header styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cell styles
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Daily Check-ins Sheet
    ws_checkins = wb.active
    ws_checkins.title = "Daily Check-ins"

    # Header styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cell styles
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws_checkins.merge_cells('A1:J1')
    title_cell = ws_checkins['A1']
    title_cell.value = f"Weekly Progress Report - Client {client.client_serial}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = header_alignment

    # Week info
    ws_checkins.merge_cells('A2:J2')
    week_cell = ws_checkins['A2']
    week_cell.value = f"Week {week_num}, {year} ({week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')})"
    week_cell.font = Font(size=14)
    week_cell.alignment = header_alignment

    # Headers
    headers = ['Date', 'Day', 'Check-in Time', 'Emotional (1-5)', 'Emotional Notes',
               'Medication', 'Medication Notes', 'Activity (1-5)', 'Activity Notes', 'Completion']

    for col, header in enumerate(headers, 1):
        cell = ws_checkins.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Get check-ins for the week
    checkins = client.checkins.filter(
        DailyCheckin.checkin_date.between(week_start.date(), week_end.date())
    ).order_by(DailyCheckin.checkin_date).all()

    # Color fills for ratings
    excellent_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green
    good_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Yellow
    poor_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Red

    # Populate daily check-ins
    row = 5
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    checkins_completed = 0

    # Variables for summary statistics
    emotional_values = []
    medication_values = []
    activity_values = []

    for i in range(7):
        current_date = week_start + timedelta(days=i)
        checkin = next((c for c in checkins if c.checkin_date == current_date.date()), None)

        ws_checkins.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
        ws_checkins.cell(row=row, column=2).value = days[i]

        if checkin:
            checkins_completed += 1
            ws_checkins.cell(row=row, column=3).value = checkin.checkin_time.strftime('%H:%M')

            # Emotional rating with color coding
            emotional_cell = ws_checkins.cell(row=row, column=4)
            emotional_cell.value = checkin.emotional_value
            if checkin.emotional_value:
                emotional_values.append(checkin.emotional_value)
                if checkin.emotional_value >= 4:
                    emotional_cell.fill = excellent_fill
                elif checkin.emotional_value == 3:
                    emotional_cell.fill = good_fill
                else:
                    emotional_cell.fill = poor_fill

            ws_checkins.cell(row=row, column=5).value = checkin.emotional_notes or ''

            # Medication formatting
            med_cell = ws_checkins.cell(row=row, column=6)
            if checkin.medication_value == 0:
                med_cell.value = "N/A"
                med_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            elif checkin.medication_value == 1:
                med_cell.value = "No"
                med_cell.fill = poor_fill
                medication_values.append(checkin.medication_value)
            elif checkin.medication_value == 3:
                med_cell.value = "Partial"
                med_cell.fill = good_fill
                medication_values.append(checkin.medication_value)
            elif checkin.medication_value == 5:
                med_cell.value = "Yes"
                med_cell.fill = excellent_fill
                medication_values.append(checkin.medication_value)

            ws_checkins.cell(row=row, column=7).value = checkin.medication_notes or ''

            # Activity rating with color coding
            activity_cell = ws_checkins.cell(row=row, column=8)
            activity_cell.value = checkin.activity_value
            if checkin.activity_value:
                activity_values.append(checkin.activity_value)
                if checkin.activity_value >= 4:
                    activity_cell.fill = excellent_fill
                elif checkin.activity_value == 3:
                    activity_cell.fill = good_fill
                else:
                    activity_cell.fill = poor_fill

            ws_checkins.cell(row=row, column=9).value = checkin.activity_notes or ''
            ws_checkins.cell(row=row, column=10).value = "✓"
            ws_checkins.cell(row=row, column=10).fill = excellent_fill
        else:
            ws_checkins.cell(row=row, column=3).value = "No check-in"
            ws_checkins.cell(row=row, column=3).font = Font(italic=True, color="999999")
            ws_checkins.cell(row=row, column=10).value = "✗"
            ws_checkins.cell(row=row, column=10).fill = poor_fill

        # Apply borders to all cells
        for col in range(1, 11):
            ws_checkins.cell(row=row, column=col).border = cell_border

        row += 1

    # 2. Weekly Summary Sheet
    ws_summary = wb.create_sheet("Weekly Summary")

    # Summary title
    ws_summary.merge_cells('A1:E1')
    summary_title = ws_summary['A1']
    summary_title.value = "Weekly Summary Statistics"
    summary_title.font = Font(bold=True, size=16)
    summary_title.alignment = header_alignment

    # Summary headers
    summary_headers = ['Metric', 'Value', 'Percentage', 'Rating', 'Notes']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Calculate statistics
    summary_data = []

    # Check-in completion
    completion_rate = (checkins_completed / 7) * 100
    summary_data.append({
        'metric': 'Check-in Completion',
        'value': f"{checkins_completed}/7 days",
        'percentage': f"{completion_rate:.1f}%",
        'rating': 'Excellent' if completion_rate >= 80 else 'Good' if completion_rate >= 60 else 'Needs Improvement',
        'notes': ''
    })

    if checkins_completed > 0:
        # Average emotional rating
        if emotional_values:
            avg_emotional = sum(emotional_values) / len(emotional_values)
            summary_data.append({
                'metric': 'Average Emotional Rating',
                'value': f"{avg_emotional:.2f}/5",
                'percentage': f"{(avg_emotional / 5) * 100:.1f}%",
                'rating': 'Excellent' if avg_emotional >= 4 else 'Good' if avg_emotional >= 3 else 'Needs Support',
                'notes': ''
            })

        # Medication adherence
        if medication_values:
            med_adherent = sum(1 for val in medication_values if val == 5)
            adherence_rate = (med_adherent / len(medication_values)) * 100
            summary_data.append({
                'metric': 'Medication Adherence',
                'value': f"{med_adherent}/{len(medication_values)} days",
                'percentage': f"{adherence_rate:.1f}%",
                'rating': 'Excellent' if adherence_rate >= 90 else 'Good' if adherence_rate >= 70 else 'Needs Improvement',
                'notes': ''
            })

        # Average activity
        if activity_values:
            avg_activity = sum(activity_values) / len(activity_values)
            summary_data.append({
                'metric': 'Average Activity Level',
                'value': f"{avg_activity:.2f}/5",
                'percentage': f"{(avg_activity / 5) * 100:.1f}%",
                'rating': 'Excellent' if avg_activity >= 4 else 'Good' if avg_activity >= 3 else 'Needs Encouragement',
                'notes': ''
            })

    # Write summary data
    row = 4
    for data in summary_data:
        ws_summary.cell(row=row, column=1).value = data['metric']
        ws_summary.cell(row=row, column=2).value = data['value']
        ws_summary.cell(row=row, column=3).value = data['percentage']

        rating_cell = ws_summary.cell(row=row, column=4)
        rating_cell.value = data['rating']
        if 'Excellent' in data['rating']:
            rating_cell.fill = excellent_fill
        elif 'Good' in data['rating']:
            rating_cell.fill = good_fill
        else:
            rating_cell.fill = poor_fill

        ws_summary.cell(row=row, column=5).value = data['notes']

        # Apply borders
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).border = cell_border

        row += 1

    # 3. Weekly Goals Sheet
    ws_goals = wb.create_sheet("Weekly Goals")

    # Goals title
    ws_goals.merge_cells('A1:I1')
    goals_title = ws_goals['A1']
    goals_title.value = "Weekly Goals & Completion"
    goals_title.font = Font(bold=True, size=16)
    goals_title.alignment = header_alignment

    # Goals headers
    goal_headers = ['Goal', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun', 'Completion Rate']
    for col, header in enumerate(goal_headers[0:1], 1):
        cell = ws_goals.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    for col, header in enumerate(goal_headers[1:8], 2):
        cell = ws_goals.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    cell = ws_goals.cell(row=3, column=9)
    cell.value = goal_headers[8]
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = cell_border

    # Get weekly goals
    weekly_goals = client.goals.filter_by(
        week_start=week_start.date(),
        is_active=True
    ).all()

    row = 4
    for goal in weekly_goals:
        ws_goals.cell(row=row, column=1).value = goal.goal_text

        # Get completions for each day
        completions = goal.completions.filter(
            GoalCompletion.completion_date.between(week_start.date(), week_end.date())
        ).all()

        completed_days = 0
        for day_idx in range(7):
            current_date = week_start.date() + timedelta(days=day_idx)
            completion = next((c for c in completions if c.completion_date == current_date), None)

            cell = ws_goals.cell(row=row, column=day_idx + 2)
            if completion:
                if completion.completed:
                    cell.value = "✓"
                    cell.fill = excellent_fill
                    completed_days += 1
                else:
                    cell.value = "✗"
                    cell.fill = poor_fill
            else:
                cell.value = "-"
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            cell.alignment = Alignment(horizontal="center")
            cell.border = cell_border

        # Completion rate
        completion_rate = (completed_days / 7) * 100
        rate_cell = ws_goals.cell(row=row, column=9)
        rate_cell.value = f"{completed_days}/7 ({completion_rate:.0f}%)"
        if completion_rate >= 80:
            rate_cell.fill = excellent_fill
        elif completion_rate >= 50:
            rate_cell.fill = good_fill
        else:
            rate_cell.fill = poor_fill
        rate_cell.border = cell_border

        # Apply borders to goal text
        ws_goals.cell(row=row, column=1).border = cell_border

        row += 1

    # 4. Therapist Notes Sheet
    ws_notes = wb.create_sheet("Therapist Notes")

    # Notes title
    ws_notes.merge_cells('A1:D1')
    notes_title = ws_notes['A1']
    notes_title.value = "Therapist Notes & Missions"
    notes_title.font = Font(bold=True, size=16)
    notes_title.alignment = header_alignment

    # Notes headers
    note_headers = ['Date', 'Type', 'Content', 'Status']
    for col, header in enumerate(note_headers, 1):
        cell = ws_notes.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Get therapist notes for the week
    week_start_datetime = datetime.combine(week_start.date(), datetime.min.time())
    week_end_datetime = datetime.combine(week_end.date(), datetime.max.time())

    notes = TherapistNote.query.filter(
        TherapistNote.client_id == client.id,
        TherapistNote.therapist_id == therapist.id,
        TherapistNote.created_at.between(week_start_datetime, week_end_datetime)
    ).order_by(TherapistNote.created_at).all()

    row = 4
    for note in notes:
        ws_notes.cell(row=row, column=1).value = note.created_at.strftime('%Y-%m-%d %H:%M')

        type_cell = ws_notes.cell(row=row, column=2)
        if note.is_mission:
            type_cell.value = "MISSION"
            type_cell.font = Font(bold=True, color="E91E63")
        else:
            type_cell.value = note.note_type.title()

        ws_notes.cell(row=row, column=3).value = note.content

        status_cell = ws_notes.cell(row=row, column=4)
        if note.is_mission:
            if note.mission_completed:
                status_cell.value = "Completed"
                status_cell.fill = excellent_fill
            else:
                status_cell.value = "Pending"
                status_cell.fill = good_fill
        else:
            status_cell.value = "-"

        # Apply borders
        for col in range(1, 5):
            ws_notes.cell(row=row, column=col).border = cell_border

        row += 1

    # 5. Additional Tracking Categories Sheet (if any)
    additional_categories = []
    for plan in client.tracking_plans.filter_by(is_active=True):
        if plan.category.name not in ['Emotion Level', 'Medication', 'Physical Activity']:
            additional_categories.append(plan.category)

    if additional_categories:
        ws_tracking = wb.create_sheet("Additional Tracking")

        # Title
        ws_tracking.merge_cells('A1:I1')
        tracking_title = ws_tracking['A1']
        tracking_title.value = "Additional Tracking Categories"
        tracking_title.font = Font(bold=True, size=16)
        tracking_title.alignment = header_alignment

        # Headers
        tracking_headers = ['Date', 'Day'] + [cat.name for cat in additional_categories]
        for col, header in enumerate(tracking_headers, 1):
            cell = ws_tracking.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Data
        row = 4
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            ws_tracking.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
            ws_tracking.cell(row=row, column=2).value = days[i]

            # Get responses for each category
            for col_idx, category in enumerate(additional_categories, 3):
                response = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    category_id=category.id,
                    response_date=current_date.date()
                ).first()

                cell = ws_tracking.cell(row=row, column=col_idx)
                if response:
                    cell.value = response.value
                    if response.value >= 4:
                        cell.fill = excellent_fill
                    elif response.value == 3:
                        cell.fill = good_fill
                    else:
                        cell.fill = poor_fill
                else:
                    cell.value = "-"

                cell.alignment = Alignment(horizontal="center")
                cell.border = cell_border

            # Apply borders
            for col in range(1, len(tracking_headers) + 1):
                ws_tracking.cell(row=row, column=col).border = cell_border

            row += 1

    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass

            # Only set width if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    return wb


# ============= REPORT GENERATION =============

@app.route('/api/reports/generate/<int:client_id>/<week>', methods=['GET'])
@require_auth(['therapist'])
def generate_report(client_id, week):
    """Generate comprehensive weekly Excel report"""
    try:
        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Create Excel workbook using the shared function
        wb = create_weekly_report_excel(client, therapist, week_start, week_end, week_num, year)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        filename = f"therapy_report_{client.client_serial}_week_{week_num}_{year}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/email-report', methods=['POST'])
@require_auth(['therapist'])
def email_therapy_report():
    """Send therapy report via email"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        week = data.get('week')
        recipient_email = data.get('recipient_email')

        # Validate input
        if not client_id:
            return jsonify({'error': 'Client ID is required'}), 400
        
        if not week:
            return jsonify({'error': 'Week is required'}), 400

        # Ensure client_id is an integer
        try:
            client_id = int(client_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid client ID'}), 400

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week
        try:
            year, week_num = week.split('-W')
            year = int(year)
            week_num = int(week_num)
        except (ValueError, AttributeError):
            return jsonify({'error': 'Invalid week format'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get check-ins for summary
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start.date(), week_end.date())
        ).order_by(DailyCheckin.checkin_date).all()

        checkins_completed = len(checkins)

        # Calculate summary statistics
        avg_emotional = 0
        adherence_rate = 0
        avg_activity = 0

        if checkins_completed > 0:
            emotional_values = [c.emotional_value for c in checkins if c.emotional_value]
            if emotional_values:
                avg_emotional = sum(emotional_values) / len(emotional_values)

            med_values = [c.medication_value for c in checkins if c.medication_value and c.medication_value > 0]
            if med_values:
                med_adherent = sum(1 for val in med_values if val == 5)
                adherence_rate = (med_adherent / len(med_values)) * 100

            activity_values = [c.activity_value for c in checkins if c.activity_value]
            if activity_values:
                avg_activity = sum(activity_values) / len(activity_values)

        # Prepare email content
        email_content = f"""
Dear Colleague,

Please find attached the weekly therapy report for client {client.client_serial}.

Report Period: {week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')} (Week {week_num}, {year})

Summary:
- Check-ins completed: {checkins_completed}/7 days
- Average emotional rating: {avg_emotional:.1f}/5
- Medication adherence: {adherence_rate:.0f}%
- Average activity level: {avg_activity:.1f}/5

The attached Excel file contains:
- Detailed daily check-ins with color-coded ratings
- Weekly summary statistics
- Goal completion tracking
- Therapist notes and missions

Please review the attached report and contact me if you have any questions.

Best regards,
{therapist.name}
{therapist.organization or ''}
        """

        # If no email configuration, return content for manual sending
        if not app.config.get('MAIL_USERNAME'):
            return jsonify({
                'success': True,
                'email_content': email_content.strip(),
                'recipient': recipient_email or therapist.user.email or 'Your email',
                'subject': f"Weekly Therapy Report - Client {client.client_serial} - Week {week_num}, {year}",
                'note': 'Email configuration not set up. Please copy this content and attach the downloaded Excel file to send manually.'
            })

        # If email is configured, send it
        try:
            # Create the Excel workbook using the shared function
            wb = create_weekly_report_excel(client, therapist, week_start, week_end, week_num, year)

            # Save to BytesIO for email attachment
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Create email
            msg = MIMEMultipart()
            msg['From'] = app.config['MAIL_USERNAME']
            msg['To'] = recipient_email or therapist.user.email
            msg['Subject'] = f"Weekly Therapy Report - Client {client.client_serial} - Week {week_num}, {year}"

            # Email body
            msg.attach(MIMEText(email_content, 'plain'))

            # Attach Excel file
            excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_attachment.set_payload(excel_buffer.read())
            encoders.encode_base64(excel_attachment)
            excel_attachment.add_header(
                'Content-Disposition',
                f'attachment; filename=therapy_report_{client.client_serial}_week_{week_num}_{year}.xlsx'
            )
            msg.attach(excel_attachment)

            # Send email
            server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
            server.starttls()
            server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
            server.send_message(msg)
            server.quit()

            return jsonify({
                'success': True,
                'message': f'Report sent successfully to {recipient_email or therapist.user.email}'
            })

        except Exception as e:
            # Return the email content even if sending fails
            return jsonify({
                'success': True,
                'email_content': email_content.strip(),
                'recipient': recipient_email or therapist.user.email or 'Your email',
                'subject': f"Weekly Therapy Report - Client {client.client_serial} - Week {week_num}, {year}",
                'error': f'Failed to send email: {str(e)}',
                'note': 'Email could not be sent automatically. Please copy this content and send it manually.'
            })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= HEALTH CHECK =============

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat()
    })


# ============= INITIALIZATION =============

# Flag to ensure single initialization
_initialized = False


def initialize_database():
    """Initialize database with default data"""
    global _initialized
    if _initialized:
        return

    _initialized = True

    try:
        db.create_all()

        # Add default tracking categories if not exist
        if TrackingCategory.query.count() == 0:
            default_categories = [
                ('Emotion Level', 'Overall emotional state', True),
                ('Energy', 'Physical and mental energy levels', True),
                ('Social Activity', 'Engagement in social interactions', True),
                ('Sleep Quality', 'Quality of sleep', False),
                ('Anxiety Level', 'Level of anxiety experienced', False),
                ('Motivation', 'Level of motivation and drive', False)
            ]

            for name, description, is_default in default_categories:
                category = TrackingCategory(
                    name=name,
                    description=description,
                    is_default=is_default
                )
                db.session.add(category)

            db.session.commit()
            print("Database initialized with default tracking categories")
    except Exception as e:
        print(f"Database initialization error: {e}")
        _initialized = False


# Initialize on first request (replaces @app.before_first_request)
@app.before_request
def before_request():
    initialize_database()


# Don't initialize on import for production
# Let init_db.py handle it during deployment
if not os.environ.get('PRODUCTION'):
    with app.app_context():
        initialize_database()

if __name__ == '__main__':
    # Ensure database is created when running directly
    with app.app_context():
        db.create_all()
        initialize_database()

    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=not os.environ.get('PRODUCTION'))
